# webapp.py - Flask Web Application for RMKCET Parent Connect
"""
Complete Flask web application replacing the Streamlit UI.
Serves HTML templates with a dark glass-morphism theme.
"""
import os
import io
import csv
import json
import hashlib
import uuid
from datetime import datetime
from functools import wraps
from urllib.parse import urlparse, parse_qs
import re

from flask import (
    Flask, render_template, request, redirect, url_for,
    session, flash, jsonify, send_file, Response
)
from fpdf import FPDF

import database as db
from config import (
    SECRET_KEY, APP_NAME, APP_VERSION, DATA_DIR,
    MESSAGE_TEMPLATE, COUNTRY_CODE, DEPT_REG_PATTERNS
)

# ---------------------------------------------------------------------------
# App setup
# ---------------------------------------------------------------------------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
FRONTEND_DIR = os.path.join(os.path.dirname(BASE_DIR), "frontend")
FRONTEND_ASSETS_DIR = os.path.join(FRONTEND_DIR, "assets")
STATIC_ASSETS_DIR = os.path.join(FRONTEND_DIR, "static", "assets")

app = Flask(
    __name__,
    template_folder=os.path.join(FRONTEND_DIR, "templates"),
    static_folder=os.path.join(FRONTEND_DIR, "static"),
)
app.secret_key = SECRET_KEY
app.config["MAX_CONTENT_LENGTH"] = 16 * 1024 * 1024  # 16 MB upload limit

db.init_database()


# ---------------------------------------------------------------------------
# Auth helpers - Industrial-grade session validation
# ---------------------------------------------------------------------------
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_email" not in session:
            return redirect(url_for("login"))
        
        sid = session.get("session_id", "")
        
        # Use industrial-grade session validation
        is_valid, reason, user_email = db.validate_session_strict(
            sid, 
            request.remote_addr, 
            request.user_agent.string
        )
        
        if not is_valid:
            session.clear()
            # Provide user-friendly messages based on reason
            if "session_timeout" in reason:
                flash("Your session has expired due to inactivity. Please log in again.", "warning")
            elif "session_inactive" in reason:
                if "new_login" in reason or "new_device" in reason:
                    flash("You have been logged out because you logged in from another device.", "warning")
                elif "admin_action" in reason:
                    flash("An administrator has logged you out.", "warning")
                else:
                    flash("Your session is no longer valid. Please log in again.", "error")
            elif reason == "user_deactivated":
                flash("Your account has been deactivated. Contact an administrator.", "error")
            elif reason == "user_locked":
                flash("Your account is locked. Contact an administrator.", "error")
            else:
                flash("Session expired. Please log in again.", "error")
            return redirect(url_for("login"))
        
        # Update session activity (heartbeat)
        db.touch_session(sid)
        return f(*args, **kwargs)
    return decorated


def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if session.get("role") not in {"admin", "chief_admin"}:
            flash("Admin access required.", "error")
            return redirect(url_for("index"))

        user = db.get_user(session.get("user_email", "")) or {}
        department = (user.get("department") or "").strip()
        if session.get("role") == "chief_admin" and department and not db.is_department_active(department):
            flash("Department is currently blocked by system administration. Access is read-only limited.", "warning")
        return f(*args, **kwargs)
    return decorated


def _is_system_admin(role):
    return str(role or "").strip().lower() == "admin"


def _is_chief_admin(role):
    return str(role or "").strip().lower() == "chief_admin"


def _is_admin_portal_user(role):
    return _is_system_admin(role) or _is_chief_admin(role)


def _get_actor_scope_pairs(actor_email, actor_role):
    if _is_system_admin(actor_role):
        return None
    scopes = db.get_chief_admin_scopes(actor_email)
    return {(str(s.get("department") or "").upper(), int(s.get("year_level") or 1)) for s in scopes}


def _can_chief_admin_touch_user(actor_email, target_user):
    if not target_user:
        return False
    
    target_role = target_user.get("role")
    
    # Chief admin can manage counselors in their scope
    if target_role == "counselor":
        scopes = _get_actor_scope_pairs(actor_email, "chief_admin") or set()
        key = (str(target_user.get("department") or "").upper(), int(target_user.get("year_level") or 1))
        return key in scopes
    
    # Chief admin can manage other chief admins if they have scope overlap
    if target_role == "chief_admin":
        actor_scopes = _get_actor_scope_pairs(actor_email, "chief_admin") or set()
        target_scopes = db.get_chief_admin_scopes(target_user.get("email"))
        target_scopes_set = {(str(s.get("department") or "").upper(), int(s.get("year_level") or 1)) for s in target_scopes}
        return bool(actor_scopes & target_scopes_set)  # Check for scope intersection
    
    return False


def _get_allowed_counselor_emails_for_actor(actor_email, actor_role):
    if _is_system_admin(actor_role):
        return None
    users = db.get_scoped_users_for_admin(actor_email, actor_role)
    return [u.get("email") for u in users if u.get("role") == "counselor"]


def _can_manage_department_year(actor_email, actor_role, department, year_level):
    if _is_system_admin(actor_role):
        return True
    if not _is_chief_admin(actor_role):
        return False
    scopes = _get_actor_scope_pairs(actor_email, actor_role) or set()
    return (str(department or "").strip().upper(), int(year_level or 1)) in scopes


def _is_counselor_department_blocked(user_email, role):
    if str(role or "") != "counselor":
        return False
    user = db.get_user(user_email) or {}
    return not db.is_department_active(user.get("department"))


def _filter_activity_for_actor(activity_rows, actor_email, actor_role):
    scopes = _get_actor_scope_pairs(actor_email, actor_role)
    if scopes is None:
        return activity_rows
    filtered = []
    for row in activity_rows:
        u = db.get_user(row.get("email")) or {}
        key = (str(row.get("department") or "").upper(), int(u.get("year_level") or 1))
        if key in scopes:
            filtered.append(row)
    return filtered


def _get_admin_tab(default_tab="users"):
    """Resolve active admin tab from form/query/referrer."""
    tab = (request.form.get("tab") or request.args.get("tab") or "").strip()
    if tab:
        return tab

    ref = (request.referrer or "").strip()
    if ref:
        try:
            parsed = urlparse(ref)
            ref_tab = parse_qs(parsed.query).get("tab", [""])[0].strip()
            if ref_tab:
                return ref_tab
        except Exception:
            pass

    return default_tab


def _redirect_admin_back(default_tab="users", **extra_query):
    params = {"tab": _get_admin_tab(default_tab)}
    for key, value in extra_query.items():
        if value is None:
            continue
        sval = str(value).strip()
        if sval:
            params[key] = sval
    return redirect(url_for("admin", **params))


def _get_message_filters_from_request():
    msg_day = (request.values.get("msg_day") or "").strip()
    msg_q = (request.values.get("msg_q") or "").strip()
    msg_year = (request.values.get("msg_year") or "").strip()
    msg_month = (request.values.get("msg_month") or "").strip()
    msg_day_num = (request.values.get("msg_day_num") or "").strip()
    return msg_day, msg_q, msg_year, msg_month, msg_day_num


def _message_export_filename(ext):
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"message_activity_{stamp}.{ext}"


def _resolve_asset_file(filename):
    for root in (FRONTEND_ASSETS_DIR, STATIC_ASSETS_DIR):
        candidate = os.path.join(root, filename)
        if os.path.isfile(candidate):
            return candidate
    return None


# ---------------------------------------------------------------------------
# Context processor – inject common vars into every template
# ---------------------------------------------------------------------------
@app.context_processor
def inject_globals():
    return {
        "app_name": APP_NAME,
        "app_version": APP_VERSION,
        "current_user_email": session.get("user_email"),
        "current_user_name": session.get("user_name"),
        "current_role": session.get("role"),
        "now": datetime.now(),
    }


@app.route("/documentation/download", methods=["GET"])
@login_required
def download_role_documentation():
    role = str(session.get("role") or "counselor").strip().lower()
    doc_map = {
        "admin": "doc_admin.pdf",
        "chief_admin": "doc_chief_admin.pdf",
        "counselor": "doc_counsellor.pdf",
    }
    filename = doc_map.get(role, "doc_counsellor.pdf")
    target = _resolve_asset_file(filename)
    if not target:
        flash("Documentation file is not available yet.", "error")
        return redirect(request.referrer or url_for("index"))
    return send_file(target, as_attachment=True, download_name=filename)


@app.route("/support/templates/student", methods=["GET"])
@login_required
def download_student_template():
    filename = "student.xlsx"
    target = _resolve_asset_file(filename)
    if not target:
        flash("Student template is not available yet.", "error")
        return redirect(request.referrer or url_for("index"))
    return send_file(target, as_attachment=True, download_name=filename)


@app.route("/support/templates/marksheet", methods=["GET"])
@login_required
def download_marksheet_template():
    filename = "marksheet.xlsx"
    target = _resolve_asset_file(filename)
    if not target:
        flash("Marksheet template is not available yet.", "error")
        return redirect(request.referrer or url_for("index"))
    return send_file(target, as_attachment=True, download_name=filename)


def _normalize_metric_key(key):
    return re.sub(r"[^a-z0-9]", "", str(key or "").lower())


def _is_unknown_metric_field(raw_key, key_norm):
    raw = str(raw_key or "").strip().lower()
    if not raw:
        return True
    if raw.startswith("unnamed"):
        return True
    if re.match(r"^subject[_\s-]*\d+$", raw):
        return True
    if key_norm.startswith("unnamed"):
        return True
    return False


def _is_absent_mark(value):
    s = str(value or "").strip().lower()
    return s in {"absent", "ab", "a", "na", "-", "not attended"}


def _build_parent_subjects_table(marks, ordered_fields=None):
    """Build standardized marks block with optional caller-defined ordering."""
    if not isinstance(marks, dict):
        return ""

    attendance_values = []
    gpa_values = []
    failed_values = []
    not_attended_values = []
    subject_rows = []

    attendance_keys = {"attendance", "att"}
    gpa_keys = {"gpa", "cgpa"}
    failed_keys = {"noofsubjectsfailed", "failedsubjects", "failedcount", "nooffailedsubjects"}
    not_attended_keys = {"examnotattended", "notattended", "absentcount", "noofsubjectsabsent"}
    ignored_keys = {
        "regno", "registernumber", "name", "studentname", "department", "section",
        "batch", "semester", "test", "total", "overall", "percentage", "grade",
        "result", "status", "parentphone", "phone", "parentemail", "email",
        "sno", "slno", "serialno", "serialnumber", "rollno",
        "absentees", "absentee", "absentstudents"
    }

    for raw_key, raw_val in marks.items():
        key_norm = _normalize_metric_key(raw_key)
        value = str(raw_val or "").strip()
        if _is_unknown_metric_field(raw_key, key_norm):
            continue
        if not key_norm or key_norm in ignored_keys:
            continue

        if key_norm in attendance_keys:
            attendance_values.append(value)
            continue
        if key_norm in gpa_keys:
            gpa_values.append(value)
            continue
        if key_norm in failed_keys:
            failed_values.append(value)
            continue
        if key_norm in not_attended_keys:
            not_attended_values.append(value)
            continue

        subject_rows.append((str(raw_key).strip(), value))

    metric_values = {
        "attendance": attendance_values,
        "failed_subjects": failed_values,
        "not_attended": not_attended_values,
        "gpa": gpa_values,
    }
    metric_labels = {
        "attendance": "Attendance",
        "failed_subjects": "Failed Subjects",
        "not_attended": "Not Attended",
        "gpa": "GPA",
    }

    used_subject_idx = set()
    ordered_lines = []

    def _pop_metric(metric_key):
        values = metric_values.get(metric_key) or []
        if not values:
            return None
        return values.pop(0)

    def _match_subject(raw_key, normalized_key):
        raw_key = str(raw_key or "").strip().lower()
        normalized_key = str(normalized_key or "").strip().lower()
        for idx, (label, value) in enumerate(subject_rows):
            if idx in used_subject_idx:
                continue
            label_norm = _normalize_metric_key(label)
            if raw_key and label.lower() == raw_key:
                used_subject_idx.add(idx)
                return label, value
            if normalized_key and label_norm == normalized_key:
                used_subject_idx.add(idx)
                return label, value
        return None

    if isinstance(ordered_fields, list):
        for item in ordered_fields:
            if not isinstance(item, dict):
                continue
            item_type = str(item.get("type") or "").strip().lower()
            if item_type == "metric":
                metric_key = str(item.get("key") or "").strip().lower()
                value = _pop_metric(metric_key)
                if value is not None:
                    ordered_lines.append(f"{metric_labels.get(metric_key, metric_key)} : {value}")
            elif item_type == "subject":
                subject = _match_subject(item.get("raw_key"), item.get("normalized_key"))
                if subject:
                    ordered_lines.append(f"{subject[0]} : {subject[1]}")

    if isinstance(ordered_fields, list):
        for idx, (subject, value) in enumerate(subject_rows):
            if idx not in used_subject_idx:
                ordered_lines.append(f"{subject} : {value}")

        for key in ("attendance", "failed_subjects", "not_attended", "gpa"):
            values = metric_values.get(key) or []
            for value in values:
                ordered_lines.append(f"{metric_labels[key]} : {value}")
    else:
        for value in attendance_values:
            ordered_lines.append(f"{metric_labels['attendance']} : {value}")
        for subject, value in subject_rows:
            ordered_lines.append(f"{subject} : {value}")
        for key in ("failed_subjects", "not_attended", "gpa"):
            for value in metric_values.get(key) or []:
                ordered_lines.append(f"{metric_labels[key]} : {value}")

    lines = []
    lines.extend(ordered_lines)

    return "\n".join(lines)


def _build_parent_message(test_name, reg_no, student_name, marks):
    marks_table = _build_parent_subjects_table(marks)
    return (
        f"Dear Parent , The Following is the {test_name} Marks Secured in each Course by your son/daughter\n\n"
        f"REGISTER NUMBER :  {reg_no}\n"
        f"NAME : {student_name}\n\n"
        f"{marks_table}\n\n"
        f"Regards\n"
        f"PRINCIPAL\n"
        f"RMKCET"
    )


# ============================= PAGE ROUTES =================================

@app.route("/")
def index():
    if "user_email" in session:
        return redirect(url_for("admin" if _is_admin_portal_user(session.get("role")) else "counselor_page"))
    return redirect(url_for("login"))


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        identifier = request.form.get("identifier", "").strip()  # Can be email or name
        password = request.form.get("password", "")
        force_logout = request.form.get("force_logout") == "true"

        user = db.authenticate_user(identifier, password)
        if not user:
            flash("Invalid email/name or password.", "error")
            return render_template("login.html")

        email = user["email"]
        allowed, msg = db.check_user_access(email)
        if not allowed:
            flash(msg, "error")
            return render_template("login.html")

        # Check for existing active session on another device
        if not force_logout:
            existing_session = db.get_user_active_session(email)
            if existing_session:
                # Return with session conflict info
                return render_template("login.html", 
                    session_conflict=True,
                    existing_session=existing_session,
                    stored_identifier=identifier,
                    stored_password=password)

        # Force logout existing sessions if requested
        if force_logout:
            db.force_logout_by_email(email, "new_device_login")

        sid = str(uuid.uuid4())
        ok, msg = db.register_session(sid, email, request.remote_addr, request.user_agent.string)
        if not ok:
            flash(msg, "error")
            return render_template("login.html")

        session["user_email"] = email
        session["user_name"] = user["name"]
        session["role"] = user["role"]
        session["session_id"] = sid
        session["department"] = user.get("department", "")

        return redirect(url_for("admin" if _is_admin_portal_user(user["role"]) else "counselor_page"))

    return render_template("login.html")


@app.route("/logout")
def logout():
    sid = session.get("session_id")
    if sid:
        db.end_session(sid)
    session.clear()
    return redirect(url_for("login"))


# ======================== ADMIN PAGE =======================================

@app.route("/admin")
@login_required
@admin_required
def admin():
    actor_email = session.get("user_email")
    actor_role = session.get("role")
    current_tab = (request.args.get("tab") or "users").strip()

    if _is_chief_admin(actor_role) and current_tab in {"monitoring", "config"}:
        flash("Only system admin can access this section.", "warning")
        return redirect(url_for("admin", tab="reports"))

    allowed_scopes = _get_actor_scope_pairs(actor_email, actor_role)
    users = db.get_scoped_users_for_admin(actor_email, actor_role)
    departments = db.get_departments_for_admin(actor_email, actor_role, active_only=False)
    active_sessions = db.get_active_sessions() if _is_admin_portal_user(actor_role) else []
    # For chief admins, filter sessions to only show scoped users
    if _is_chief_admin(actor_role):
        scoped_emails = {u.get("email") for u in users}
        active_sessions = [s for s in active_sessions if s.get("user_email") in scoped_emails]
    # Create a set of emails with active sessions for status display
    logged_in_users = {s.get("user_email") for s in active_sessions if s.get("is_active") and not s.get("forced_logout")}
    full_activity = db.get_counselor_activity_summary()
    activity = _filter_activity_for_actor(full_activity, actor_email, actor_role)
    format_settings = db.get_format_settings()

    msg_day = (request.args.get("msg_day") or "").strip()
    msg_q = (request.args.get("msg_q") or "").strip()
    msg_year = (request.args.get("msg_year") or "").strip()
    msg_month = (request.args.get("msg_month") or "").strip()
    msg_day_num = (request.args.get("msg_day_num") or "").strip()
    allowed_counselors = _get_allowed_counselor_emails_for_actor(actor_email, actor_role)

    messages = db.get_message_history_filtered(
        day=msg_day or None,
        counselor_query=msg_q or None,
        limit=1500,
        filter_year=msg_year or None,
        filter_month=msg_month or None,
        filter_day=msg_day_num or None,
        allowed_counselors=allowed_counselors,
    )
    message_days = db.get_message_days(counselor_query=msg_q or None)
    grouped_map = {}
    for m in messages:
        day_key = str((m.get("sent_at") or "")[:10] or "Unknown")
        grouped_map.setdefault(day_key, []).append(m)
    message_groups = [{"day": day, "messages": rows, "total": len(rows)} for day, rows in grouped_map.items()]
    msg_stats = db.get_message_stats(actor_email) if _is_chief_admin(actor_role) else db.get_message_stats()
    counselor_suggestions = db.get_message_counselor_suggestions(
        query=msg_q,
        allowed_counselors=allowed_counselors,
    )
    
    # App configuration
    app_config = db.get_app_config()
    
    # Session monitoring statistics
    session_stats = db.get_session_statistics() if _is_system_admin(actor_role) else {
        "active_sessions": 0, "today_sessions": 0, "avg_duration_minutes": 0, "forced_logouts": 0,
        "peak_concurrent": 0, "desktop_sessions": 0, "mobile_sessions": 0, "logout_reasons": {}
    }
    session_history = db.get_session_history(limit=100) if _is_system_admin(actor_role) else []

    selected_department = (request.args.get("dept") or "").strip().upper()
    selected_year_level = request.args.get("year_level", type=int) or 1
    department_tests = db.get_all_unique_tests(
        filter_dept=selected_department or None,
        filter_year_level=selected_year_level,
        allowed_scopes=allowed_scopes,
    )

    report_department = (request.args.get("report_dept") or selected_department or "").strip().upper()
    report_year_level = request.args.get("report_year", type=int) or selected_year_level
    report_tests = db.get_all_unique_tests(
        filter_dept=report_department or None,
        filter_year_level=report_year_level,
        allowed_scopes=allowed_scopes,
    )
    recent_report_tests = db.get_all_unique_tests(allowed_scopes=allowed_scopes)[:6]
    chief_scopes = db.get_chief_admin_scopes(actor_email) if _is_chief_admin(actor_role) else []
    chief_scope_keys = [
        f"{str(s.get('department') or '').upper()}::{int(s.get('year_level') or 1)}"
        for s in chief_scopes
    ]
    chief_scopes_by_email = {}
    for u in users:
        if u.get("role") != "chief_admin":
            continue
        scopes_for_user = db.get_chief_admin_scopes(u.get("email"))
        chief_scopes_by_email[u.get("email")] = [
            f"{str(s.get('department') or '').upper()}::{int(s.get('year_level') or 1)}"
            for s in scopes_for_user
        ]

    counselors = [u for u in users if u["role"] == "counselor"]
    students_map = {c["email"]: db.get_students(c["email"]) for c in counselors}
    return render_template(
        "admin.html",
        users=users,
        departments=departments,
        sessions=active_sessions,
        activity=activity,
        format_settings=format_settings,
        messages=messages,
        message_groups=message_groups,
        message_days=message_days,
        selected_message_day=msg_day,
        message_query=msg_q,
        selected_message_year=msg_year,
        selected_message_month=msg_month,
        selected_message_day_num=msg_day_num,
        counselor_suggestions=counselor_suggestions,
        msg_stats=msg_stats,
        counselor_count=len(counselors),
        active_counselor_count=sum(1 for c in counselors if c["is_active"]),
        session_count=len(active_sessions),
        department_tests=department_tests,
        selected_department=selected_department,
        selected_year_level=selected_year_level,
        report_department=report_department,
        report_year_level=report_year_level,
        report_tests=report_tests,
        recent_report_tests=recent_report_tests,
        chief_scopes=chief_scopes,
        chief_scope_keys=chief_scope_keys,
        chief_scopes_by_email=chief_scopes_by_email,
        logged_in_users=logged_in_users,
        is_system_admin=_is_system_admin(actor_role),
        is_chief_admin=_is_chief_admin(actor_role),
        app_config=app_config,
        session_stats=session_stats,
        session_history=session_history,
        students_map=students_map,
    )


@app.route("/api/messages/delete/<int:message_id>", methods=["POST"])
@login_required
@admin_required
def api_delete_message(message_id):
    msg_day, msg_q, msg_year, msg_month, msg_day_num = _get_message_filters_from_request()
    deleted = db.delete_message_by_id(message_id)
    if deleted:
        flash("Message entry deleted.", "success")
    else:
        flash("Message entry not found.", "warning")
    return _redirect_admin_back("messages", msg_day=msg_day, msg_q=msg_q, msg_year=msg_year, msg_month=msg_month, msg_day_num=msg_day_num)


@app.route("/api/messages/delete-bulk", methods=["POST"])
@login_required
@admin_required
def api_delete_messages_bulk():
    msg_day, msg_q, msg_year, msg_month, msg_day_num = _get_message_filters_from_request()
    ids = request.form.getlist("message_ids")
    if not ids:
        flash("Select at least one message to delete.", "warning")
        return _redirect_admin_back("messages", msg_day=msg_day, msg_q=msg_q, msg_year=msg_year, msg_month=msg_month, msg_day_num=msg_day_num)

    deleted = db.delete_messages_by_ids(ids)
    if deleted:
        flash(f"Deleted {deleted} message entr{'y' if deleted == 1 else 'ies'}.", "success")
    else:
        flash("No messages were deleted.", "warning")
    return _redirect_admin_back("messages", msg_day=msg_day, msg_q=msg_q, msg_year=msg_year, msg_month=msg_month, msg_day_num=msg_day_num)


@app.route("/api/messages/export/csv")
@login_required
@admin_required
def api_export_messages_csv():
    msg_day, msg_q, msg_year, msg_month, msg_day_num = _get_message_filters_from_request()
    allowed_counselors = _get_allowed_counselor_emails_for_actor(session.get("user_email"), session.get("role"))
    data = db.get_message_history_filtered(
        day=msg_day or None,
        counselor_query=msg_q or None,
        limit=None,
        filter_year=msg_year or None,
        filter_month=msg_month or None,
        filter_day=msg_day_num or None,
        allowed_counselors=allowed_counselors,
    )

    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["Date", "Time", "Counselor", "Counselor Email", "Student", "Reg No", "Format", "Status"])
    for m in data:
        sent_at = str(m.get("sent_at") or "")
        date_part = sent_at[:10]
        time_part = sent_at[11:19] if len(sent_at) >= 19 else ""
        w.writerow([
            date_part,
            time_part,
            m.get("counselor_name") or m.get("counselor_email") or "",
            m.get("counselor_email") or "",
            m.get("student_name") or "",
            m.get("reg_no") or "",
            m.get("format") or "",
            m.get("status") or "",
        ])

    buf.seek(0)
    return Response(
        buf.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment;filename={_message_export_filename('csv')}"},
    )


@app.route("/api/messages/export/excel")
@login_required
@admin_required
def api_export_messages_excel():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    msg_day, msg_q, msg_year, msg_month, msg_day_num = _get_message_filters_from_request()
    allowed_counselors = _get_allowed_counselor_emails_for_actor(session.get("user_email"), session.get("role"))
    data = db.get_message_history_filtered(
        day=msg_day or None,
        counselor_query=msg_q or None,
        limit=None,
        filter_year=msg_year or None,
        filter_month=msg_month or None,
        filter_day=msg_day_num or None,
        allowed_counselors=allowed_counselors,
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Message Activity"

    headers = ["Date", "Time", "Counselor", "Counselor Email", "Student", "Reg No", "Format", "Status"]
    header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for ri, m in enumerate(data, 2):
        sent_at = str(m.get("sent_at") or "")
        ws.cell(row=ri, column=1, value=sent_at[:10])
        ws.cell(row=ri, column=2, value=sent_at[11:19] if len(sent_at) >= 19 else "")
        ws.cell(row=ri, column=3, value=m.get("counselor_name") or m.get("counselor_email") or "")
        ws.cell(row=ri, column=4, value=m.get("counselor_email") or "")
        ws.cell(row=ri, column=5, value=m.get("student_name") or "")
        ws.cell(row=ri, column=6, value=m.get("reg_no") or "")
        ws.cell(row=ri, column=7, value=m.get("format") or "")
        ws.cell(row=ri, column=8, value=m.get("status") or "")

    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col) + 2
        ws.column_dimensions[col[0].column_letter].width = min(max_len, 36)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=_message_export_filename("xlsx"),
    )


@app.route("/api/messages/export/pdf")
@login_required
@admin_required
def api_export_messages_pdf():
    msg_day, msg_q, msg_year, msg_month, msg_day_num = _get_message_filters_from_request()
    allowed_counselors = _get_allowed_counselor_emails_for_actor(session.get("user_email"), session.get("role"))
    data = db.get_message_history_filtered(
        day=msg_day or None,
        counselor_query=msg_q or None,
        limit=None,
        filter_year=msg_year or None,
        filter_month=msg_month or None,
        filter_day=msg_day_num or None,
        allowed_counselors=allowed_counselors,
    )

    pdf = FPDF("L")
    pdf.add_page()
    pdf.set_font("Arial", "B", 15)
    pdf.cell(0, 10, "RMKCET Parent Connect - Message Activity", 0, 1, "C")
    pdf.set_font("Arial", "", 9)
    subtitle = f"Day: {msg_day or 'All'}   Counselor Search: {msg_q or 'All'}   Generated: {datetime.now().strftime('%d-%b-%Y %H:%M')}"
    pdf.cell(0, 8, subtitle[:150], 0, 1, "C")
    pdf.ln(2)

    widths = [24, 18, 44, 52, 36, 24, 20, 20]
    heads = ["Date", "Time", "Counselor", "Email", "Student", "Reg No", "Format", "Status"]

    pdf.set_font("Arial", "B", 8)
    pdf.set_fill_color(102, 126, 234)
    pdf.set_text_color(255, 255, 255)
    for w, h in zip(widths, heads):
        pdf.cell(w, 8, h, 1, 0, "C", True)
    pdf.ln()

    pdf.set_text_color(0, 0, 0)
    pdf.set_font("Arial", "", 7)
    for m in data:
        sent_at = str(m.get("sent_at") or "")
        pdf.cell(widths[0], 7, sent_at[:10], 1)
        pdf.cell(widths[1], 7, (sent_at[11:19] if len(sent_at) >= 19 else ""), 1)
        pdf.cell(widths[2], 7, str(m.get("counselor_name") or m.get("counselor_email") or "")[:24], 1)
        pdf.cell(widths[3], 7, str(m.get("counselor_email") or "")[:30], 1)
        pdf.cell(widths[4], 7, str(m.get("student_name") or "")[:20], 1)
        pdf.cell(widths[5], 7, str(m.get("reg_no") or "")[:14], 1)
        pdf.cell(widths[6], 7, str(m.get("format") or "")[:10], 1, 0, "C")
        pdf.cell(widths[7], 7, str(m.get("status") or "")[:10], 1, 1, "C")

    buf = io.BytesIO()
    raw = pdf.output(dest="S")
    buf.write(raw if isinstance(raw, bytes) else raw.encode("latin-1"))
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/pdf",
        as_attachment=True,
        download_name=_message_export_filename("pdf"),
    )


# ======================== COUNSELOR PAGE ===================================

@app.route("/counselor")
@login_required
def counselor_page():
    email = session["user_email"]
    user = db.get_user(email)
    if not user:
        session.clear()
        return redirect(url_for("login"))

    is_blocked_department = not db.is_department_active(user.get("department"))
    students = db.get_students(email)
    tests = db.get_visible_tests_for_counselor(email)
    recent_tests = tests[:2]
    msg_stats = db.get_message_stats(email)
    msg_history = db.get_message_history(email, limit=50)
    
    # Get counselor's previous marksheet submissions
    submissions = db.get_counselor_submissions(email, limit=50)

    # Send-flow state (selected test and pending students)
    selected_test_id = request.args.get("test_id", type=int)
    valid_ids = {int(t.get("id")) for t in tests if t.get("id") is not None}
    if selected_test_id not in valid_ids:
        selected_test_id = int(tests[0]["id"]) if tests else None

    selected_test_meta = db.get_test_metadata(selected_test_id) if selected_test_id else None
    pending_students = db.get_pending_students_for_test(email, selected_test_id) if selected_test_id else students
    sent_reg_nos = db.get_sent_reg_nos_for_test(email, selected_test_id) if selected_test_id else set()

    return render_template(
        "counselor.html",
        user=user,
        is_blocked_department=is_blocked_department,
        students=students,
        assigned_students_count=len(students),
        tests=tests,
        recent_tests=recent_tests,
        msg_stats=msg_stats,
        msg_history=msg_history,
        submissions=submissions,
        selected_test_id=selected_test_id,
        selected_test_meta=selected_test_meta,
        pending_students=pending_students,
        sent_count=(len(sent_reg_nos) if selected_test_id else 0),
        can_upload_students=bool(user.get("can_upload_students", 1)),
        report_tab=(request.args.get("tab") or "recent-tests"),
    )


# ============================== API ROUTES ==================================

# ---------- Users -----------------------------------------------------------

@app.route("/api/users", methods=["POST"])
@login_required
@admin_required
def api_create_user():
    actor_email = session.get("user_email")
    actor_role = session.get("role")
    email = request.form.get("email", "").strip()
    password = request.form.get("password", "")
    confirm_password = request.form.get("confirm_password", "")
    name = request.form.get("name", "").strip()
    role = request.form.get("role", "counselor")
    role = role if role in {"admin", "chief_admin", "counselor"} else "counselor"
    year_level = request.form.get("year_level", type=int) or 1
    scope_values = request.form.getlist("chief_scopes")

    if _is_chief_admin(actor_role) and role != "counselor":
        flash("Chief admins can create counselor accounts only.", "error")
        return _redirect_admin_back("users")

    if role == "admin" and not _is_system_admin(actor_role):
        flash("Only system admin can create system admin accounts.", "error")
        return _redirect_admin_back("users")

    if role == "admin":
        # Admins get unrestricted defaults and do not need counselor-specific fields.
        department = ""
        year_level = 1
        max_students = 500
        can_upload = True
    elif role == "chief_admin":
        department = (request.form.get("department") or "").strip().upper()
        year_level = 1
        max_students = 500
        can_upload = True
    else:
        department = request.form.get("department", "").strip()
        max_students_raw = request.form.get("max_students", "30")
        can_upload = request.form.get("can_upload_students") == "on"

        try:
            max_students = int(max_students_raw)
        except (TypeError, ValueError):
            flash("Max students must be a valid number.", "error")
            return _redirect_admin_back("users")

        if max_students < 1 or max_students > 500:
            flash("Max students must be between 1 and 500.", "error")
            return _redirect_admin_back("users")

    if not email or not password or not name:
        flash("All required fields must be filled.", "error")
        return _redirect_admin_back("users")

    if password != confirm_password:
        flash("Password and confirm password do not match.", "error")
        return _redirect_admin_back("users")

    if len(password) < 6:
        flash("Password must be at least 6 characters.", "error")
        return _redirect_admin_back("users")

    if role == "counselor" and year_level not in (1, 2, 3, 4):
        flash("Year must be between 1 and 4.", "error")
        return _redirect_admin_back("users")

    if _is_chief_admin(actor_role) and role == "counselor":
        scopes = _get_actor_scope_pairs(actor_email, actor_role) or set()
        if (str(department or "").upper(), int(year_level or 1)) not in scopes:
            flash("You can only create counselors inside your assigned department/year scope.", "error")
            return _redirect_admin_back("users")

    ok, msg = db.create_user(
        email, password, name, role, department, max_students, can_upload, year_level
    )

    if ok and role == "chief_admin":
        parsed_scopes = []
        for raw in scope_values:
            dep, _, yr = str(raw or "").partition("::")
            try:
                yr_int = int(yr)
            except (TypeError, ValueError):
                continue
            parsed_scopes.append((dep, yr_int))
        if parsed_scopes:
            db.set_chief_admin_scopes(email, parsed_scopes)

    # Optional student file during registration (counselors only)
    if ok and role == "counselor" and "student_file" in request.files:
        f = request.files["student_file"]
        if f and f.filename:
            try:
                from core.dynamic_parser import parse_student_excel
                parsed = parse_student_excel(f)
                if parsed:
                    added = db.add_students_bulk(email, parsed)
                    flash(f"User created — {added} students uploaded.", "success")
                else:
                    flash("User created but no valid students found in the uploaded file.", "warning")
            except Exception as e:
                flash(f"User created but student upload failed: {e}", "warning")
            return _redirect_admin_back("users")

    flash(msg, "success" if ok else "error")
    return _redirect_admin_back("users")


@app.route("/api/users/<path:email>/update", methods=["POST"])
@login_required
@admin_required
def api_update_user(email):
    actor_email = session.get("user_email")
    actor_role = session.get("role")
    target = db.get_user(email)
    if not target:
        flash("User not found.", "error")
        return _redirect_admin_back("users")

    if _is_chief_admin(actor_role):
        if target.get("role") != "counselor":
            flash("You can modify only counselor accounts.", "error")
            return _redirect_admin_back("users")

        actor_scopes = _get_actor_scope_pairs(actor_email, actor_role) or set()
        target_key = (
            str(target.get("department") or "").strip().upper(),
            int(target.get("year_level") or 1),
        )
        if target_key not in actor_scopes:
            flash("You can modify only counselors in your assigned scope.", "error")
            return _redirect_admin_back("users")

        requested_dep = (request.form.get("department") or target.get("department") or "").strip().upper()
        requested_year = request.form.get("year_level", type=int) or int(target.get("year_level") or 1)
        requested_key = (requested_dep, requested_year)
        if requested_key not in actor_scopes:
            flash("Update rejected: target department/year is outside your authorized assignments.", "error")
            return _redirect_admin_back("users")

    updates = {}
    name = request.form.get("name", "").strip()
    if name:
        updates["name"] = name

    requested_role = str(target.get("role") or "counselor").strip().lower() or "counselor"
    if _is_system_admin(actor_role):
        role_input = str(request.form.get("role") or requested_role).strip().lower()
        if role_input not in {"admin", "chief_admin", "counselor"}:
            flash("Invalid role selected.", "error")
            return _redirect_admin_back("users")
        requested_role = role_input
        updates["role"] = requested_role

    password = request.form.get("password", "").strip()
    if password:
        if len(password) < 6:
            flash("Password must be at least 6 characters.", "error")
            return _redirect_admin_back("users")
        updates["password"] = password

    if _is_system_admin(actor_role):
        if requested_role == "admin":
            updates["department"] = ""
            updates["year_level"] = 1
            updates["max_students"] = 500
            updates["can_upload_students"] = 1
        elif requested_role == "chief_admin":
            department = (request.form.get("department") or "").strip().upper()
            updates["department"] = department
            updates["year_level"] = 1
            updates["max_students"] = 500
            updates["can_upload_students"] = 1
        else:
            department = (request.form.get("department") or "").strip().upper()
            if not department:
                flash("Department is required for counselor accounts.", "error")
                return _redirect_admin_back("users")
            year_level = request.form.get("year_level", type=int)
            if year_level not in (1, 2, 3, 4):
                flash("Year must be between 1 and 4.", "error")
                return _redirect_admin_back("users")

            max_students_raw = request.form.get("max_students", "")
            try:
                max_students = int(max_students_raw)
            except (TypeError, ValueError):
                flash("Max students must be a valid number.", "error")
                return _redirect_admin_back("users")

            if max_students < 1 or max_students > 500:
                flash("Max students must be between 1 and 500.", "error")
                return _redirect_admin_back("users")

            updates["department"] = department
            updates["year_level"] = year_level
            updates["max_students"] = max_students
            updates["can_upload_students"] = 1 if request.form.get("can_upload_students") == "on" else 0
    else:
        department = (request.form.get("department") or target.get("department") or "").strip().upper()
        year_level = request.form.get("year_level", type=int) or int(target.get("year_level") or 1)
        requested_key = (department, year_level)

        actor_scopes = _get_actor_scope_pairs(actor_email, actor_role) or set()
        if requested_key not in actor_scopes:
            flash("Update rejected: target department/year is outside your authorized assignments.", "error")
            return _redirect_admin_back("users")

        max_students_raw = request.form.get("max_students")
        if max_students_raw:
            try:
                max_students = int(max_students_raw)
            except (TypeError, ValueError):
                flash("Max students must be a valid number.", "error")
                return _redirect_admin_back("users")
            if max_students < 1 or max_students > 500:
                flash("Max students must be between 1 and 500.", "error")
                return _redirect_admin_back("users")
            updates["max_students"] = max_students

        updates["can_upload_students"] = 1 if request.form.get("can_upload_students") == "on" else 0

    db.update_user(email, **updates)

    if _is_system_admin(actor_role):
        if requested_role == "chief_admin":
            scope_values = request.form.getlist("chief_scopes")
            parsed_scopes = []
            for raw in scope_values:
                dep, _, yr = str(raw or "").partition("::")
                try:
                    parsed_scopes.append((dep, int(yr)))
                except (TypeError, ValueError):
                    continue
            db.set_chief_admin_scopes(email, parsed_scopes)
        else:
            db.set_chief_admin_scopes(email, [])

    flash("User updated.", "success")
    return _redirect_admin_back("users")


@app.route("/api/users/<path:email>/delete", methods=["POST"])
@login_required
@admin_required
def api_delete_user(email):
    actor_email = session.get("user_email")
    actor_role = session.get("role")
    target = db.get_user(email)
    if not target:
        flash("User not found.", "error")
        return _redirect_admin_back("users")
    if _is_chief_admin(actor_role) and not _can_chief_admin_touch_user(actor_email, target):
        flash("You can delete only counselors in your assigned scope.", "error")
        return _redirect_admin_back("users")
    db.delete_user(email)
    flash("User deleted.", "success")
    return _redirect_admin_back("users")


@app.route("/api/users/<path:email>/lock", methods=["POST"])
@login_required
@admin_required
def api_lock_user(email):
    actor_email = session.get("user_email")
    actor_role = session.get("role")
    target = db.get_user(email)
    if _is_chief_admin(actor_role) and not _can_chief_admin_touch_user(actor_email, target):
        flash("You can lock only counselors in your assigned scope.", "error")
        return _redirect_admin_back("users")
    db.lock_user(email, request.form.get("reason", "Locked by admin"))
    flash("User locked.", "success")
    return _redirect_admin_back("users")


@app.route("/api/users/<path:email>/unlock", methods=["POST"])
@login_required
@admin_required
def api_unlock_user(email):
    actor_email = session.get("user_email")
    actor_role = session.get("role")
    target = db.get_user(email)
    if _is_chief_admin(actor_role) and not _can_chief_admin_touch_user(actor_email, target):
        flash("You can unlock only counselors in your assigned scope.", "error")
        return _redirect_admin_back("users")
    db.unlock_user(email)
    flash("User unlocked.", "success")
    return _redirect_admin_back("users")


@app.route("/api/password-update", methods=["POST"])
@login_required
def api_update_password():
    """Update password for logged in user."""
    user_email = session.get("user_email")
    if not user_email:
        flash("Session expired. Please login again.", "error")
        return redirect(url_for("login"))
    
    current_password = request.form.get("current_password", "").strip()
    new_password = request.form.get("new_password", "").strip()
    confirm_password = request.form.get("confirm_password", "").strip()
    
    # Validate inputs
    if not current_password or not new_password or not confirm_password:
        flash("All password fields are required.", "error")
        return redirect(request.referrer or url_for("admin"))
    
    if len(new_password) < 6:
        flash("New password must be at least 6 characters.", "error")
        return redirect(request.referrer or url_for("admin"))
    
    if new_password != confirm_password:
        flash("Passwords do not match.", "error")
        return redirect(request.referrer or url_for("admin"))
    
    # Verify current password
    user = db.get_user(user_email) or {}
    stored_hash = user.get("password_hash")
    
    if not stored_hash or not hashlib.sha256(current_password.encode()).hexdigest() == stored_hash:
        flash("Current password is incorrect.", "error")
        return redirect(request.referrer or url_for("admin"))
    
    # Update password
    new_hash = hashlib.sha256(new_password.encode()).hexdigest()
    db.update_user_password(user_email, new_hash)
    
    flash("Password updated successfully.", "success")
    return redirect(request.referrer or url_for("admin"))


@app.route("/api/users/<path:email>/upload-students", methods=["POST"])
@login_required
@admin_required
def api_upload_students_for_counselor(email):
    actor_email = session.get("user_email")
    actor_role = session.get("role")
    target = db.get_user(email)
    if _is_chief_admin(actor_role) and not _can_chief_admin_touch_user(actor_email, target):
        flash("You can upload students only for counselors in your assigned scope.", "error")
        return _redirect_admin_back("users")
    f = request.files.get("student_file")
    if not f or not f.filename:
        flash("No file selected.", "error")
        return _redirect_admin_back("users")
    try:
        from core.dynamic_parser import parse_student_excel
        parsed = parse_student_excel(f)
        if parsed:
            added = db.add_students_bulk(email, parsed)
            flash(f"{added} students uploaded for {email}.", "success")
        else:
            flash("No valid students found.", "error")
    except Exception as e:
        flash(f"Upload failed: {e}", "error")
    return _redirect_admin_back("users")


@app.route("/api/users/<path:email>/force-logout", methods=["POST"])
@login_required
@admin_required
def api_force_logout(email):
    actor_email = session.get("user_email")
    actor_role = session.get("role")
    target = db.get_user(email)
    if _is_chief_admin(actor_role) and not _can_chief_admin_touch_user(actor_email, target):
        flash("You can force logout only counselors in your assigned scope.", "error")
        return _redirect_admin_back("users")
    db.force_logout_user(email, "admin_action")
    flash(f"Force-logged-out {email}.", "success")
    return _redirect_admin_back("users")


@app.route("/api/users/reset-password", methods=["POST"])
@login_required
@admin_required
def api_admin_reset_password():
    actor_email = session.get("user_email")
    actor_role = session.get("role")
    target_email = request.form.get("target_email", "").strip()
    new_password = request.form.get("new_password", "")
    confirm_password = request.form.get("confirm_password", "")
    force_logout = request.form.get("force_logout") == "on"

    if not target_email or not new_password or not confirm_password:
        flash("User and both password fields are required.", "error")
        return redirect(url_for("admin", tab="config"))

    if new_password != confirm_password:
        flash("New password and confirm password do not match.", "error")
        return redirect(url_for("admin", tab="config"))

    if len(new_password) < 6:
        flash("Password must be at least 6 characters.", "error")
        return redirect(url_for("admin", tab="config"))

    user = db.get_user(target_email)
    if not user:
        flash("Selected user was not found.", "error")
        return redirect(url_for("admin", tab="config"))

    if _is_chief_admin(actor_role) and not _can_chief_admin_touch_user(actor_email, user):
        flash("You can reset passwords only for counselors in your assigned scope.", "error")
        return redirect(url_for("admin", tab="users"))

    db.update_user(target_email, password=new_password)

    # Security best-practice: invalidate existing sessions after password reset.
    if force_logout and target_email != session.get("user_email"):
        db.force_logout_user(target_email, "admin_password_reset")

    flash(f"Password updated successfully for {target_email}.", "success")
    return redirect(url_for("admin", tab="config"))


@app.route("/api/chief-admin/reset-counselor-password", methods=["POST"])
@login_required
@admin_required
def api_chief_admin_reset_password():
    """Chief admin can reset password for counselors in their assigned dept/year scope."""
    actor_email = session.get("user_email")
    actor_role = session.get("role")
    
    if not _is_chief_admin(actor_role):
        flash("Only chief admins can access this function.", "error")
        return redirect(url_for("counselor_page"))
    
    target_email = request.form.get("target_email", "").strip()
    new_password = request.form.get("new_password", "")
    confirm_password = request.form.get("confirm_password", "")
    force_logout = request.form.get("force_logout") == "on"

    if not target_email or not new_password or not confirm_password:
        flash("Counselor and both password fields are required.", "error")
        return redirect(url_for("counselor_page"))

    if new_password != confirm_password:
        flash("New password and confirm password do not match.", "error")
        return redirect(url_for("counselor_page"))

    if len(new_password) < 6:
        flash("Password must be at least 6 characters.", "error")
        return redirect(url_for("counselor_page"))

    target = db.get_user(target_email)
    if not target:
        flash("Selected user was not found.", "error")
        return redirect(url_for("counselor_page"))
    
    if target.get("role") != "counselor":
        flash("You can reset passwords only for counselors.", "error")
        return redirect(url_for("counselor_page"))

    # Chief admin must be able to manage this counselor's department/year
    if not _can_chief_admin_touch_user(actor_email, target):
        flash("You can reset passwords only for counselors in your assigned scope.", "error")
        return redirect(url_for("counselor_page"))

    db.update_user(target_email, password=new_password)

    # Security best-practice: invalidate existing sessions after password reset.
    if force_logout and target_email != session.get("user_email"):
        db.force_logout_user(target_email, "chief_admin_password_reset")

    flash(f"Password updated successfully for {target_email}.", "success")
    return redirect(url_for("counselor_page"))


@app.route("/api/chief-admin/scoped-counselors", methods=["GET"])
@login_required
@admin_required
def api_chief_admin_get_scoped_counselors():
    """Return list of counselors under chief admin's dept/year scope as JSON."""
    actor_email = session.get("user_email")
    actor_role = session.get("role")
    
    if not _is_chief_admin(actor_role):
        return jsonify({"error": "Unauthorized"}), 403
    
    # Get all users scoped to this chief admin
    all_scoped_users = db.get_scoped_users_for_admin(actor_email, actor_role) or []
    
    # Filter to only counselors
    counselors = [
        {
            "name": u.get("name", ""),
            "email": u.get("email", ""),
            "department": u.get("department", ""),
            "year_level": u.get("year_level", 1)
        }
        for u in all_scoped_users
        if u.get("role") == "counselor"
    ]
    
    return jsonify(counselors)


@app.route("/api/admin/students/save", methods=["POST"])
@login_required
@admin_required
def api_admin_save_student():
    counselor_email = request.form.get("counselor_email", "").strip()
    original_reg_no = request.form.get("original_reg_no", "").strip()
    reg_no = request.form.get("reg_no", "").strip()
    student_name = request.form.get("student_name", "").strip()
    parent_phone = request.form.get("parent_phone", "").strip()
    parent_email = request.form.get("parent_email", "").strip()

    if not counselor_email or not reg_no or not student_name:
        flash("Counselor, Register No and Student Name are required.", "error")
        return _redirect_admin_back("users", open_manage=counselor_email)

    counselor = db.get_user(counselor_email)
    if not counselor or counselor.get("role") != "counselor":
        flash("Invalid counselor selected.", "error")
        return _redirect_admin_back("users")

    department = (counselor.get("department") or "").strip().upper()

    try:
        if original_reg_no and original_reg_no != reg_no:
            db.delete_student(counselor_email, original_reg_no)

        db.admin_upsert_student(
            counselor_email,
            reg_no,
            student_name,
            department=department,
            parent_phone=parent_phone,
            parent_email=parent_email,
        )
        flash(f"Student {reg_no} saved for {counselor.get('name')}", "success")
    except Exception as e:
        flash(f"Could not save student: {e}", "error")

    return _redirect_admin_back("users", open_manage=counselor_email)


@app.route("/api/admin/students/delete", methods=["POST"])
@login_required
@admin_required
def api_admin_delete_student():
    counselor_email = request.form.get("counselor_email", "").strip()
    reg_no = request.form.get("reg_no", "").strip()

    if not counselor_email or not reg_no:
        flash("Counselor and Register No are required.", "error")
        return _redirect_admin_back("users", open_manage=counselor_email)

    try:
        db.delete_student(counselor_email, reg_no)
        flash(f"Deleted student {reg_no}.", "success")
    except Exception as e:
        flash(f"Could not delete student: {e}", "error")

    return _redirect_admin_back("users", open_manage=counselor_email)


@app.route("/api/admin/students/delete-all", methods=["POST"])
@login_required
@admin_required
def api_admin_delete_all_students():
    counselor_email = request.form.get("counselor_email", "").strip()
    if not counselor_email:
        flash("Counselor is required.", "error")
        return _redirect_admin_back("users")

    counselor = db.get_user(counselor_email)
    if not counselor or counselor.get("role") != "counselor":
        flash("Invalid counselor selected.", "error")
        return _redirect_admin_back("users")

    try:
        db.delete_all_students(counselor_email)
        flash(f"Deleted all students for {counselor.get('name')}", "success")
    except Exception as e:
        flash(f"Could not delete student list: {e}", "error")

    return _redirect_admin_back("users", open_manage=counselor_email)


# ---------- Departments -----------------------------------------------------

@app.route("/api/departments", methods=["POST"])
@login_required
@admin_required
def api_create_department():
    if _is_chief_admin(session.get("role")):
        flash("Only system admin can create departments.", "error")
        return _redirect_admin_back("departments")
    code = request.form.get("code", "").strip().upper()
    name = request.form.get("name", "").strip()
    color = request.form.get("color", "#667eea")
    if not code or not name:
        flash("Code and name are required.", "error")
        return _redirect_admin_back("departments")
    ok, msg = db.create_department(code, name, color)
    flash(msg, "success" if ok else "error")
    return _redirect_admin_back("departments")


@app.route("/api/departments/<int:dept_id>/delete", methods=["POST"])
@login_required
@admin_required
def api_delete_department(dept_id):
    if _is_chief_admin(session.get("role")):
        flash("Only system admin can delete departments.", "error")
        return _redirect_admin_back("departments")
    db.delete_department(dept_id)
    flash("Department deleted.", "success")
    return _redirect_admin_back("departments")


@app.route("/api/departments/<int:dept_id>/toggle", methods=["POST"])
@login_required
@admin_required
def api_toggle_department(dept_id):
    actor_role = session.get("role")
    if _is_chief_admin(actor_role):
        flash("Only system admin can enable or disable departments.", "error")
        return _redirect_admin_back("departments")

    is_active = request.form.get("is_active") == "1"
    db.update_department(dept_id, is_active=0 if is_active else 1)
    flash("Department updated.", "success")
    return _redirect_admin_back("departments")


# ---------- Sessions --------------------------------------------------------

@app.route("/api/sessions/cleanup", methods=["POST"])
@login_required
@admin_required
def api_cleanup_sessions():
    db.cleanup_stale_sessions()
    db.clear_inactive_sessions()
    flash("Sessions cleaned.", "success")
    return _redirect_admin_back("monitoring")


@app.route("/api/sessions/logout-all", methods=["POST"])
@login_required
@admin_required
def api_logout_all():
    db.logout_all_users()
    flash("All users logged out.", "success")
    return _redirect_admin_back("monitoring")


# ---------- App Configuration -----------------------------------------------

@app.route("/api/config/update", methods=["POST"])
@login_required
@admin_required
def api_update_config():
    settings = {}
    
    # Session timeout
    timeout = request.form.get("session_timeout")
    if timeout:
        try:
            settings["session_timeout"] = str(int(timeout))
        except ValueError:
            flash("Invalid session timeout value.", "error")
            return _redirect_admin_back("config")
    
    # Heartbeat interval
    heartbeat = request.form.get("session_heartbeat_interval")
    if heartbeat:
        try:
            settings["session_heartbeat_interval"] = str(int(heartbeat))
        except ValueError:
            pass
    
    # Hex color settings
    color_fields = [
        "color_primary", "color_primary_dark", "color_secondary", "color_accent",
        "color_success", "color_warning", "color_danger", "color_info",
        "color_bg_primary", "color_bg_secondary",
        "color_text", "color_text_dim", "color_text_muted"
    ]
    for field in color_fields:
        value = request.form.get(field)
        if value and value.startswith("#"):
            settings[field] = value

    # Advanced color fields that may use rgba()/hex formats
    advanced_color_fields = ["color_bg_card", "color_border"]
    for field in advanced_color_fields:
        value = request.form.get(field)
        if value:
            settings[field] = value.strip()
    
    # Session monitoring settings
    session_monitoring = request.form.get("session_monitoring_enabled")
    settings["session_monitoring_enabled"] = "true" if session_monitoring == "on" else "false"
    
    allow_concurrent = request.form.get("allow_concurrent_sessions")
    settings["allow_concurrent_sessions"] = "true" if allow_concurrent == "on" else "false"
    
    max_concurrent = request.form.get("max_concurrent_sessions")
    if max_concurrent:
        settings["max_concurrent_sessions"] = str(max_concurrent)
    
    if settings:
        db.update_app_config_bulk(settings)
        flash("Configuration updated successfully.", "success")
    
    return _redirect_admin_back("config")


@app.route("/api/config/reset-theme", methods=["POST"])
@login_required
@admin_required
def api_reset_theme():
    # Reset all color settings to defaults
    defaults = {
        "color_primary": "#667eea",
        "color_primary_dark": "#5a6fd6",
        "color_secondary": "#764ba2",
        "color_accent": "#a78bfa",
        "color_success": "#25D366",
        "color_warning": "#f59e0b",
        "color_danger": "#ef4444",
        "color_info": "#3b82f6",
        "color_bg_primary": "#0a0c14",
        "color_bg_secondary": "#0f1219",
        "color_bg_card": "rgba(20, 30, 50, 0.65)",
        "color_text": "#e2e8f0",
        "color_text_dim": "#94a3b8",
        "color_text_muted": "#64748b",
        "color_border": "rgba(102, 126, 234, 0.18)"
    }
    db.update_app_config_bulk(defaults)
    flash("All theme colors reset to defaults.", "success")
    return _redirect_admin_back("config")


# ---------- Activity Export -------------------------------------------------

@app.route("/api/activity/export/csv")
@login_required
@admin_required
def api_export_activity_csv():
    data = _filter_activity_for_actor(
        db.get_counselor_activity_summary(),
        session.get("user_email"),
        session.get("role"),
    )
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["Name", "Email", "Department", "Students", "Tests",
                "Messages", "Status", "Last Login"])
    for a in data:
        w.writerow([a["name"], a["email"], a["department"],
                     a["student_count"], a["tests_uploaded"],
                     a["total_messages"], a["work_status"],
                     a["last_login"] or "Never"])
    buf.seek(0)
    return Response(
        buf.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment;filename=counselor_activity.csv"},
    )


@app.route("/api/activity/export/excel")
@login_required
@admin_required
def api_export_activity_excel():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    data = _filter_activity_for_actor(
        db.get_counselor_activity_summary(),
        session.get("user_email"),
        session.get("role"),
    )
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Counselor Activity"

    headers = ["Name", "Email", "Department", "Students", "Tests",
               "Messages", "Week Msgs", "Status", "Last Login"]
    header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for ri, a in enumerate(data, 2):
        ws.cell(row=ri, column=1, value=a["name"])
        ws.cell(row=ri, column=2, value=a["email"])
        ws.cell(row=ri, column=3, value=a["department"])
        ws.cell(row=ri, column=4, value=a["student_count"])
        ws.cell(row=ri, column=5, value=a["tests_uploaded"])
        ws.cell(row=ri, column=6, value=a["total_messages"])
        ws.cell(row=ri, column=7, value=a["week_messages"])
        ws.cell(row=ri, column=8, value=a["work_status"])
        ws.cell(row=ri, column=9, value=a["last_login"] or "Never")

    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col) + 2
        ws.column_dimensions[col[0].column_letter].width = min(max_len, 35)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name="counselor_activity.xlsx")


@app.route("/api/activity/export/pdf")
@login_required
@admin_required
def api_export_activity_pdf():
    data = _filter_activity_for_actor(
        db.get_counselor_activity_summary(),
        session.get("user_email"),
        session.get("role"),
    )
    pdf = FPDF("L")
    pdf.add_page()
    pdf.set_font("Arial", "B", 16)
    pdf.cell(0, 12, "RMKCET Parent Connect - Counselor Activity", 0, 1, "C")
    pdf.set_font("Arial", "I", 10)
    pdf.cell(0, 8, f"Generated: {datetime.now().strftime('%d-%b-%Y %H:%M')}", 0, 1, "C")
    pdf.ln(4)

    widths = [45, 55, 28, 22, 20, 26, 42, 40]
    heads = ["Name", "Email", "Dept", "Students", "Tests", "Messages", "Status", "Last Login"]
    pdf.set_font("Arial", "B", 9)
    pdf.set_fill_color(102, 126, 234)
    pdf.set_text_color(255, 255, 255)
    for w, h in zip(widths, heads):
        pdf.cell(w, 8, h, 1, 0, "C", True)
    pdf.ln()

    pdf.set_text_color(0, 0, 0)
    pdf.set_font("Arial", "", 8)
    for a in data:
        login = (a["last_login"] or "Never")[:16]
        pdf.cell(widths[0], 7, a["name"][:22], 1)
        pdf.cell(widths[1], 7, a["email"][:28], 1)
        pdf.cell(widths[2], 7, a["department"][:10], 1, 0, "C")
        pdf.cell(widths[3], 7, str(a["student_count"]), 1, 0, "C")
        pdf.cell(widths[4], 7, str(a["tests_uploaded"]), 1, 0, "C")
        pdf.cell(widths[5], 7, str(a["total_messages"]), 1, 0, "C")
        pdf.cell(widths[6], 7, a["work_status"][:22], 1, 0, "C")
        pdf.cell(widths[7], 7, login, 1, 1)

    buf = io.BytesIO()
    raw = pdf.output(dest="S")
    buf.write(raw if isinstance(raw, bytes) else raw.encode("latin-1"))
    buf.seek(0)
    return send_file(buf, mimetype="application/pdf",
                     as_attachment=True, download_name="counselor_activity.pdf")


@app.route("/api/activity/<path:email>")
@login_required
@admin_required
def api_activity_detail(email):
    if _is_chief_admin(session.get("role")):
        target = db.get_user(email)
        if not _can_chief_admin_touch_user(session.get("user_email"), target):
            return jsonify({"error": "Access denied"}), 403

    detail = db.get_counselor_detailed_activity(email)
    if not detail:
        return jsonify({"error": "Not found"}), 404
    # Make JSON-serializable
    for key in list(detail.keys()):
        val = detail[key]
        if isinstance(val, datetime):
            detail[key] = val.isoformat()
    return jsonify(detail)


# ---------- Counselor student / marks / reports -----------------------------

@app.route("/api/students/upload", methods=["POST"])
@login_required
def api_upload_students():
    email = session["user_email"]
    user = db.get_user(email)
    if not user.get("can_upload_students", 1):
        flash("You do not have permission to upload students.", "error")
        return redirect(url_for("counselor_page"))

    f = request.files.get("student_file")
    if not f or not f.filename:
        flash("No file selected.", "error")
        return redirect(url_for("counselor_page"))

    try:
        from core.dynamic_parser import parse_student_excel
        parsed = parse_student_excel(f)
        if parsed:
            added = db.add_students_bulk(email, parsed)
            flash(f"{added} students uploaded successfully.", "success")
        else:
            flash("No valid students found in the file.", "error")
    except Exception as e:
        flash(f"Upload failed: {e}", "error")
    return redirect(url_for("counselor_page"))


@app.route("/api/students/<reg_no>/delete", methods=["POST"])
@login_required
def api_delete_student(reg_no):
    db.delete_student(session["user_email"], reg_no)
    flash("Student removed.", "success")
    return redirect(url_for("counselor_page"))


@app.route("/api/students/delete-all", methods=["POST"])
@login_required
def api_delete_all_students():
    db.delete_all_students(session["user_email"])
    flash("All students removed.", "success")
    return redirect(url_for("counselor_page"))


# ---------- Tests (Admin) ---------------------------------------------------

@app.route("/api/tests/<int:test_id>/delete", methods=["POST"])
@login_required
@admin_required
def api_delete_test(test_id):
    try:
        actor_email = session.get("user_email")
        actor_role = session.get("role")
        meta = db.get_test_metadata(test_id) or {}
        if not _can_manage_department_year(actor_email, actor_role, meta.get("department"), meta.get("year_level") or 1):
            flash("You can manage tests only in your assigned department/year scope.", "error")
            return _redirect_admin_back("reports")
        db.delete_test(test_id)
        flash("Test deleted successfully.", "success")
    except Exception as e:
        flash(f"Failed to delete test: {e}", "error")
    return _redirect_admin_back("reports")


@app.route("/api/tests/<int:test_id>/update", methods=["POST"])
@login_required
@admin_required
def api_update_test(test_id):
    try:
        actor_email = session.get("user_email")
        actor_role = session.get("role")
        meta = db.get_test_metadata(test_id) or {}
        if not _can_manage_department_year(actor_email, actor_role, meta.get("department"), meta.get("year_level") or 1):
            flash("You can manage tests only in your assigned department/year scope.", "error")
            return _redirect_admin_back("reports")
        test_name = (request.form.get("test_name") or "").strip() or (meta.get("test_name") or "")
        semester = (request.form.get("semester") or "").strip() or (meta.get("semester") or "")
        department = (request.form.get("department") or "").strip() or (meta.get("department") or "")
        batch_name = (request.form.get("batch_name") or "").strip() or (meta.get("batch_name") or "")
        section = (request.form.get("section") or "").strip() or (meta.get("section") or "")

        db.update_test_metadata_fields(
            test_id,
            test_name=test_name,
            semester=semester,
            department=department,
            batch_name=batch_name,
            section=section,
        )
        flash("Test updated successfully.", "success")
    except Exception as e:
        flash(f"Failed to update test: {e}", "error")
    return _redirect_admin_back("reports")


@app.route("/api/tests/<int:test_id>/toggle-block", methods=["POST"])
@login_required
@admin_required
def api_toggle_test_block(test_id):
    actor_email = session.get("user_email")
    actor_role = session.get("role")
    meta = db.get_test_metadata(test_id) or {}
    if not meta:
        flash("Test not found.", "error")
        return _redirect_admin_back("reports")

    if not _can_manage_department_year(actor_email, actor_role, meta.get("department"), meta.get("year_level") or 1):
        flash("You can manage tests only in your assigned department/year scope.", "error")
        return _redirect_admin_back("reports")

    current = int(meta.get("is_blocked") or 0)
    next_value = 0 if current else 1
    db.update_test_block_status(test_id, next_value)
    flash("Test blocked." if next_value else "Test unblocked.", "success")
    return _redirect_admin_back(
        "reports",
        report_dept=(meta.get("department") or "").strip().upper(),
        report_year=int(meta.get("year_level") or 1),
    )


@app.route("/api/admin/tests/upload", methods=["POST"])
@login_required
@admin_required
def api_admin_upload_marksheet():
    f = request.files.get("marks_file")
    if not f or not f.filename:
        flash("No marks file selected.", "error")
        return _redirect_admin_back("reports")

    department = (request.form.get("department") or "").strip().upper()
    year_level = request.form.get("year_level", type=int) or 1
    semester = (request.form.get("semester") or "").strip()
    batch_name = (request.form.get("batch_name") or "").strip()
    section = (request.form.get("section") or "").strip()
    test_name_input = (request.form.get("test_name") or "").strip()
    upload_mode = (request.form.get("upload_mode") or "new").strip().lower()

    if not department or year_level not in (1, 2, 3, 4) or not semester or not batch_name:
        flash("Department, year, semester and batch are required.", "error")
        return _redirect_admin_back("reports", report_dept=department, report_year=year_level)

    if not _can_manage_department_year(session.get("user_email"), session.get("role"), department, year_level):
        flash("You can upload tests only in your assigned department/year scope.", "error")
        return _redirect_admin_back("reports", report_dept=department, report_year=year_level)

    try:
        file_bytes = f.read()
        file_hash = hashlib.sha256(file_bytes).hexdigest()

        from core.intelligent_parser import IntelligentParser
        parser = IntelligentParser()
        test_info, students = parser.parse_file(io.BytesIO(file_bytes), f.filename)

        if not students:
            flash("No student marks data found in file.", "error")
            return _redirect_admin_back("reports", report_dept=department, report_year=year_level)

        subjects = [s["name"] for s in test_info.subjects]
        if not subjects:
            flash("Upload blocked: no subject columns detected.", "warning")
            return _redirect_admin_back("reports", report_dept=department, report_year=year_level)

        student_data = [s.to_dict() for s in students]
        test_name = test_name_input or (test_info.test_name or "Unit Test")

        existing = db.find_existing_department_year_test(
            department=department,
            year_level=year_level,
            semester=semester,
            test_name=test_name,
            batch_name=batch_name,
        )

        replace_test_id = None
        if existing:
            if (existing.get("file_hash") or "") == file_hash:
                flash("Duplicate file detected for this department/year/test. Upload blocked.", "warning")
                return _redirect_admin_back("reports", report_dept=department, report_year=year_level)
            if upload_mode == "replace":
                replace_test_id = int(existing.get("test_id"))

        ok, msg = db.save_test_marks(
            test_name=test_name,
            semester=semester,
            counselor_email=session["user_email"],
            students=student_data,
            subjects=subjects,
            batch_name=batch_name,
            department=department,
            section=section,
            file_hash=file_hash,
            replace_test_id=replace_test_id,
            sync_students=False,
            year_level=year_level,
            enforce_assigned_match=False,
            uploaded_by=session["user_email"],
        )
        if ok:
            flash(f"Marksheet uploaded for {department} Year {year_level} ({len(student_data)} students).", "success")
        else:
            flash(f"Upload failed: {msg}", "error")
    except Exception as e:
        flash(f"Upload failed: {e}", "error")

    return _redirect_admin_back("reports", report_dept=department, report_year=year_level)


@app.route("/api/tests/<int:test_id>/counselor-update", methods=["POST"])
@login_required
def api_counselor_update_test(test_id):
    email = session.get("user_email")
    role = session.get("role", "counselor")
    if _is_counselor_department_blocked(email, role):
        flash("Your department is blocked. Editing is disabled.", "warning")
        return redirect(url_for("counselor_page", tab="test-database"))
    if not _can_access_test_for_user(test_id, email, role):
        flash("Access denied for this test.", "error")
        return redirect(url_for("counselor_page", tab="test-database"))
    if _is_test_blocked(test_id):
        flash("This test is blocked by administration. Editing is disabled.", "warning")
        return redirect(url_for("counselor_page", tab="test-database"))

    try:
        meta = db.get_test_metadata(test_id) or {}
        test_name = (request.form.get("test_name") or "").strip() or (meta.get("test_name") or "")
        semester = (request.form.get("semester") or "").strip() or (meta.get("semester") or "")
        batch_name = (request.form.get("batch_name") or "").strip() or (meta.get("batch_name") or "")
        section = (request.form.get("section") or "").strip() or (meta.get("section") or "")

        db.update_test_metadata_fields(
            test_id,
            test_name=test_name,
            semester=semester,
            batch_name=batch_name,
            section=section,
        )
        flash("Test details updated.", "success")
    except Exception as e:
        flash(f"Could not update test: {e}", "error")

    return redirect(url_for("counselor_page", tab="test-database"))


@app.route("/api/tests/<int:test_id>/marks")
@login_required
def api_get_test_marks(test_id):
    """Get test marks grouped by student for display."""
    try:
        user_email = session.get("user_email")
        role = session.get("role")
        if not _can_access_test_for_user(test_id, user_email, role):
            return jsonify({"success": False, "error": "Access denied for this test."}), 403
        if role == "counselor" and _is_test_blocked(test_id):
            return jsonify({"success": False, "error": "This test is blocked by administration."}), 403

        if role == "counselor":
            data = db.get_test_marks_grouped_for_counselor(test_id, user_email)
        else:
            data = db.get_test_marks_grouped(test_id)
        return jsonify({"success": True, "data": data})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route("/api/counselor/tests/<int:test_id>/marks/update", methods=["POST"])
@login_required
def api_counselor_update_marks(test_id):
    email = session.get("user_email")
    role = session.get("role")
    if role != "counselor":
        return jsonify({"success": False, "error": "Counselor access required."}), 403
    if not _can_access_test_for_user(test_id, email, role):
        return jsonify({"success": False, "error": "Access denied."}), 403
    if _is_test_blocked(test_id):
        return jsonify({"success": False, "error": "This test is blocked by administration."}), 403
    if _is_counselor_department_blocked(email, role):
        return jsonify({"success": False, "error": "Department is blocked. Editing disabled."}), 403

    try:
        payload = request.get_json(force=True, silent=False) or {}
        reg_no = str(payload.get("reg_no") or "").strip()
        marks = payload.get("marks") or {}
        if not reg_no or not isinstance(marks, dict):
            return jsonify({"success": False, "error": "Invalid payload."}), 400

        for subject_name, value in marks.items():
            if not str(subject_name or "").strip():
                continue
            db.upsert_counselor_mark_override(email, test_id, reg_no, str(subject_name), str(value or ""))

        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/api/tests/cleanup-duplicates", methods=["POST"])
@login_required
@admin_required
def api_tests_cleanup_duplicates():
    try:
        deleted_count = db.cleanup_duplicate_tests()
        if deleted_count > 0:
            flash(f"Removed {deleted_count} duplicate test(s).", "success")
        else:
            flash("No duplicate tests found.", "info")
    except Exception as e:
        flash(f"Failed to cleanup duplicates: {e}", "error")
    return _redirect_admin_back("reports")


@app.route("/api/marks/upload", methods=["POST"])
@login_required
def api_upload_marks():
    flash("Marksheet upload is now admin-only. Use Departments tab in Admin panel.", "warning")
    if _is_admin_portal_user(session.get("role")):
        return _redirect_admin_back("reports")
    return redirect(url_for("counselor_page", tab="test-database"))

    email = session["user_email"]
    user = db.get_user(email) or {}
    f = request.files.get("marks_file")
    if not f or not f.filename:
        flash("No file selected.", "error")
        return redirect(url_for("counselor_page"))

    try:
        file_bytes = f.read()
        file_hash = hashlib.sha256(file_bytes).hexdigest()
        upload_mode = request.form.get("upload_mode", "new")
        replace_test_id = request.form.get("replace_test_id", type=int)

        from core.intelligent_parser import IntelligentParser
        parser = IntelligentParser()
        test_info, students = parser.parse_file(io.BytesIO(file_bytes), f.filename)

        if not students:
            flash("No student marks data found in file.", "error")
            return redirect(url_for("counselor_page"))

        subjects = [s["name"] for s in test_info.subjects]
        student_data = [s.to_dict() for s in students]

        if not subjects:
            flash("UPLOAD BLOCKED: HEADER INTEGRITY FAILED - no subject columns detected. FILE NOT UPLOADED.", "warning")
            return redirect(url_for("counselor_page", tab="test-database"))

        def _norm_dept(value):
            return re.sub(r"[^A-Za-z0-9]", "", str(value or "").upper())

        def _norm_reg(value):
            reg = str(value or "").strip().replace(" ", "")
            if reg.endswith(".0"):
                reg = reg[:-2]
            return reg.upper()

        def _infer_department_from_regnos(rows):
            """Best-effort department inference using configured register patterns."""
            score = {}
            for st in rows:
                reg = _norm_reg(st.get("reg_no"))
                if not reg:
                    continue
                digits = "".join(ch for ch in reg if ch.isdigit())
                if not digits:
                    continue
                for dept_code, pattern in DEPT_REG_PATTERNS.items():
                    if str(pattern) and str(pattern) in digits:
                        score[dept_code] = score.get(dept_code, 0) + 1

            if not score:
                return ""
            winner = max(score, key=score.get)
            return winner

        counselor_department = (user.get("department") or "").strip()
        parsed_department = (test_info.department or "").strip()
        if not counselor_department:
            flash("UPLOAD BLOCKED: counselor department is not configured. Contact admin. FILE NOT UPLOADED.", "warning")
            return redirect(url_for("counselor_page", tab="test-database"))

        if not parsed_department:
            parsed_department = _infer_department_from_regnos(student_data)

        if not parsed_department:
            # Header formats vary a lot across departments; fallback keeps uploads usable
            # while counselor ownership checks below still prevent cross-department data injection.
            parsed_department = counselor_department
            flash("Header department not detected. Used your counselor department for validation.", "info")

        if _norm_dept(parsed_department) != _norm_dept(counselor_department):
            flash(
                f"UPLOAD BLOCKED: NO MATCH - marksheet department '{parsed_department}' does not match your department '{counselor_department}'. FILE NOT UPLOADED.",
                "warning",
            )
            return redirect(url_for("counselor_page", tab="test-database"))

        # Enforce counselor scope: only admin-assigned students whose Reg No + Name match.
        def _norm_name(value):
            name = str(value or "").strip().lower()
            name = re.sub(r"\s+", " ", name)
            return name

        assigned_students = db.get_students(email)
        assigned_map = {
            _norm_reg(s.get("reg_no")): _norm_name(s.get("student_name"))
            for s in assigned_students
            if _norm_reg(s.get("reg_no"))
        }

        matched_students = []
        mismatch_examples = []
        mismatch_count = 0
        for st in student_data:
            reg = _norm_reg(st.get("reg_no"))
            name = _norm_name(st.get("name"))
            assigned_name = assigned_map.get(reg)
            if not reg or not assigned_name:
                mismatch_count += 1
                if len(mismatch_examples) < 5:
                    mismatch_examples.append(f"{st.get('reg_no', '')} ({st.get('name', '')})")
                continue
            if assigned_name != name:
                mismatch_count += 1
                if len(mismatch_examples) < 5:
                    mismatch_examples.append(f"{st.get('reg_no', '')} ({st.get('name', '')})")
                continue
            st["reg_no"] = reg
            matched_students.append(st)

        if not matched_students:
            sample = ", ".join(mismatch_examples) if mismatch_examples else "No valid rows"
            flash(
                f"UPLOAD BLOCKED: NO MATCH - no assigned students matched by Reg No + Name. Examples: {sample}. FILE NOT UPLOADED.",
                "warning",
            )
            return redirect(url_for("counselor_page", tab="test-database"))

        total_rows = len(student_data)
        if mismatch_count > 0:
            sample = ", ".join(mismatch_examples) if mismatch_examples else "Check uploaded rows"
            flash(
                f"PARTIAL MATCH: scanned {total_rows} rows, matched {len(matched_students)}, skipped {mismatch_count} non-assigned/mismatched rows. Examples: {sample}.",
                "warning",
            )

        student_data = matched_students

        raw_test_name = (request.form.get("test_name") or test_info.test_name or "Unit Test").strip()

        def _canonical_test_name(name):
            n = re.sub(r"\s+", " ", str(name or "").strip().lower())
            table = {
                "unit test 1": "Unit Test 1",
                "ut1": "Unit Test 1",
                "unit test 2": "Unit Test 2",
                "ut2": "Unit Test 2",
                "iat 1": "IAT 1",
                "internal assessment test 1": "IAT 1",
                "iat 2": "IAT 2",
                "internal assessment test 2": "IAT 2",
                "model": "Model",
                "model exam": "Model",
                "model to be sent": "Model",
            }
            return table.get(n)

        test_name = _canonical_test_name(raw_test_name) or raw_test_name or "Unit Test"

        semester = (request.form.get("semester") or str(test_info.semester or "1")).strip()
        batch_name = (request.form.get("batch_name") or test_info.batch_name or "").strip()
        section = (request.form.get("section") or test_info.section or "").strip()
        department = counselor_department

        existing_test = db.find_existing_department_test(department, semester, test_name, batch_name=batch_name)
        if existing_test and upload_mode != "replace":
            if (existing_test.get("file_hash") or "") == file_hash:
                flash("UPLOAD BLOCKED: DUPLICATE FILE for this department/test. FILE NOT UPLOADED.", "warning")
                return redirect(url_for("counselor_page", tab="test-database"))
            replace_test_id = int(existing_test.get("test_id"))

        if upload_mode == "replace" and not replace_test_id and existing_test:
            replace_test_id = int(existing_test.get("test_id"))

        ok, msg = db.save_test_marks(
            test_name,
            semester,
            email,
            student_data,
            subjects,
            batch_name=batch_name,
            department=department,
            section=section,
            file_hash=file_hash,
            replace_test_id=replace_test_id,
            sync_students=False,
        )
        if ok:
            # Automatic duplicate cleanup after upload.
            db.cleanup_duplicate_tests()
            latest_test_id = db.get_latest_test_id_for_counselor(email)
            verb = "updated" if replace_test_id else "uploaded"
            msg_text = f"Marks {verb} — scanned {total_rows} rows, uploaded {len(student_data)} assigned students, {len(subjects)} subjects."
            flash(msg_text, "success")
            return redirect(url_for("counselor_test_send_page", test_id=latest_test_id or replace_test_id))
        else:
            if "no match" in str(msg).lower():
                flash(msg, "warning")
            else:
                flash(f"Error: {msg}", "error")
    except Exception as e:
        flash(f"Parse error: {e}", "error")
    return redirect(url_for("counselor_page", tab="test-database"))


def _can_access_test_for_user(test_id: int, user_email: str, role: str) -> bool:
    if _is_system_admin(role):
        return True
    if _is_chief_admin(role):
        meta = db.get_test_metadata(test_id) or {}
        return _can_manage_department_year(user_email, role, meta.get("department"), meta.get("year_level") or 1)
    tests = db.get_visible_tests_for_counselor(user_email)
    allowed_ids = {int(t.get("id")) for t in tests if t.get("id") is not None}
    return test_id in allowed_ids


def _is_test_blocked(test_id: int) -> bool:
    meta = db.get_test_metadata(test_id) or {}
    return bool(int(meta.get("is_blocked") or 0))


@app.route("/counselor/tests/<int:test_id>/view")
@login_required
def counselor_test_view_page(test_id):
    email = session.get("user_email")
    role = session.get("role", "counselor")
    if not _can_access_test_for_user(test_id, email, role):
        flash("Access denied for this test.", "error")
        return redirect(url_for("counselor_page", tab="test-database"))
    if _is_test_blocked(test_id):
        flash("This test is blocked by administration. Viewing is disabled.", "warning")
        return redirect(url_for("counselor_page", tab="test-database"))
    if _is_counselor_department_blocked(email, role):
        flash("Your department is blocked. Contact system admin.", "warning")
        return redirect(url_for("counselor_page", tab="recent-tests"))

    test_meta = db.get_test_metadata(test_id) or {}
    grouped = db.get_test_marks_grouped_for_counselor(test_id, email) if role == "counselor" else db.get_test_marks_grouped(test_id)
    if role != "admin":
        def _norm_reg(value):
            reg = str(value or "").strip().replace(" ", "")
            if reg.endswith(".0"):
                reg = reg[:-2]
            return reg.upper()

        allowed_reg_nos = {_norm_reg(s.get("reg_no", "")) for s in db.get_students(email)}
        grouped["students"] = [
            s for s in grouped.get("students", [])
            if _norm_reg(s.get("reg_no", "")) in allowed_reg_nos
        ]
    return render_template(
        "counselor_test_view.html",
        test_id=test_id,
        test_meta=test_meta,
        subjects=grouped.get("subjects", []),
        students=grouped.get("students", []),
    )


@app.route("/counselor/tests/<int:test_id>/send")
@login_required
def counselor_test_send_page(test_id):
    email = session.get("user_email")
    role = session.get("role", "counselor")
    if not _can_access_test_for_user(test_id, email, role):
        flash("Access denied for this test.", "error")
        return redirect(url_for("counselor_page", tab="test-database"))
    if _is_test_blocked(test_id):
        flash("This test is blocked by administration. Sending is disabled.", "warning")
        return redirect(url_for("counselor_page", tab="test-database"))
    if _is_counselor_department_blocked(email, role):
        flash("Your department is blocked. Sending is disabled.", "warning")
        return redirect(url_for("counselor_page", tab="recent-tests"))

    user = db.get_user(email)
    students = db.get_students(email)
    def _norm_reg(value):
        reg = str(value or "").strip().replace(" ", "")
        if reg.endswith(".0"):
            reg = reg[:-2]
        return reg.upper()

    by_reg = {_norm_reg(s.get("reg_no")): s for s in students}
    grouped = db.get_test_marks_grouped_for_counselor(test_id, email) if role == "counselor" else db.get_test_marks_grouped(test_id)
    test_meta = db.get_test_metadata(test_id) or {}
    sent_reg_nos = {_norm_reg(r) for r in db.get_sent_reg_nos_for_test(email, test_id)}

    rows = []
    for sm in grouped.get("students", []):
        reg_no = sm.get("reg_no")
        norm_reg = _norm_reg(reg_no)
        if norm_reg not in by_reg:
            continue
        stu = by_reg.get(norm_reg, {})
        rows.append({
            "reg_no": norm_reg,
            "student_name": stu.get("student_name", reg_no),
            "parent_phone": stu.get("parent_phone", ""),
            "department": stu.get("department") or test_meta.get("department") or user.get("department") or "",
            "marks": sm.get("marks", {}),
            "status": "Generated" if norm_reg in sent_reg_nos else "Pending",
        })

    return render_template(
        "counselor_send_results.html",
        test_id=test_id,
        test_meta=test_meta,
        rows=rows,
        country_code=COUNTRY_CODE,
    )


@app.route("/api/reports/send-single", methods=["POST"])
@login_required
def api_send_single_report():
    email = session.get("user_email")
    role = session.get("role", "counselor")
    test_id = request.form.get("test_id", type=int)
    reg_no = request.form.get("reg_no", "").strip()
    action = request.form.get("action", "cancel")
    is_ajax = request.form.get("ajax") == "1"
    ordered_fields_raw = request.form.get("ordered_fields", "").strip()

    ordered_fields = None
    if ordered_fields_raw:
        try:
            parsed = json.loads(ordered_fields_raw)
            if isinstance(parsed, list):
                ordered_fields = parsed
        except Exception:
            ordered_fields = None

    if not test_id or not reg_no:
        flash("Test and student are required.", "error")
        return redirect(url_for("counselor_page", tab="test-database"))

    if not _can_access_test_for_user(test_id, email, role):
        flash("Access denied for this test.", "error")
        return redirect(url_for("counselor_page", tab="test-database"))
    if _is_test_blocked(test_id):
        flash("This test is blocked by administration. Sending is disabled.", "warning")
        if is_ajax:
            return jsonify({"success": False, "error": "Test is blocked."}), 403
        return redirect(url_for("counselor_page", tab="test-database"))
    if _is_counselor_department_blocked(email, role):
        flash("Your department is blocked. Sending is disabled.", "warning")
        if is_ajax:
            return jsonify({"success": False, "error": "Department is blocked."}), 403
        return redirect(url_for("counselor_page", tab="recent-tests"))

    if action != "send":
        flash("Message status kept as Pending.", "info")
        if is_ajax:
            return jsonify({"success": True, "status": "pending"})
        return redirect(url_for("counselor_test_send_page", test_id=test_id))

    user = db.get_user(email)
    test_meta = db.get_test_metadata(test_id) or {}
    students = db.get_students(email)
    def _norm_reg(value):
        reg = str(value or "").strip().replace(" ", "")
        if reg.endswith(".0"):
            reg = reg[:-2]
        return reg.upper()

    normalized_reg = _norm_reg(reg_no)
    stu = next((s for s in students if _norm_reg(s.get("reg_no")) == normalized_reg), None)
    if not stu:
        flash("Student not found under your account.", "error")
        return redirect(url_for("counselor_test_send_page", test_id=test_id))

    if role == "counselor":
        marks = db.get_student_marks_for_reg_for_counselor(test_id, normalized_reg, email)
    else:
        marks = db.get_student_marks_for_reg(test_id, normalized_reg)
    if not marks:
        flash("No marks found for selected student.", "error")
        return redirect(url_for("counselor_test_send_page", test_id=test_id))

    from utils.whatsapp_helper import get_whatsapp_link
    from utils.template_engine import TemplateEngine

    template = request.form.get("message_template", "").strip() or (
        "Dear Parent , The Following is the {test_name} Marks Secured in each Course by your son/daughter\n\n"
        "REGISTER NUMBER :  {reg_no}\n"
        "NAME : {student_name}\n\n"
        "{subjects_table}\n\n"
        "Regards\n"
        "PRINCIPAL\n"
        "RMKCET"
    )

    marks_table = _build_parent_subjects_table(marks, ordered_fields=ordered_fields)
    effective_test_name = request.form.get("test_name") or (test_meta.get("test_name") or "Unit Test")
    if request.form.get("message_template", "").strip():
        msg = TemplateEngine.fill_template(
            template,
            app_name=APP_NAME,
            reg_no=normalized_reg,
            student_name=stu.get("student_name", normalized_reg),
            department=request.form.get("department") or (test_meta.get("department") or stu.get("department", "")),
            test_name=effective_test_name,
            semester=request.form.get("semester") or (test_meta.get("semester") or "-"),
            batch_name=request.form.get("batch_name") or (test_meta.get("batch_name") or "-"),
            section=request.form.get("section") or (test_meta.get("section") or "-"),
            subjects_table=marks_table,
            counselor_name=user.get("name", "Counselor"),
        )
    else:
        msg = _build_parent_message(
            effective_test_name,
            normalized_reg,
            stu.get("student_name", normalized_reg),
            marks,
        )

    def _clean_phone(value):
        digits = "".join(ch for ch in str(value or "") if ch.isdigit())
        return digits[-10:] if len(digits) >= 10 else ""

    def _clean_reg(value):
        reg = str(value or "").strip().replace(" ", "")
        if reg.endswith(".0"):
            reg = reg[:-2]
        return reg

    phone = _clean_phone(stu.get("parent_phone", ""))
    if not phone:
        # Fallback for older uploads where phone may have been parsed into email-like field.
        fallback_phone = _clean_phone(stu.get("parent_email", ""))
        if fallback_phone:
            phone = fallback_phone
            try:
                db.update_student(email, reg_no, parent_phone=phone)
            except Exception:
                pass

    if not phone:
        # Secondary fallback: recover from any duplicate/legacy student row with equivalent reg number.
        target_reg = _clean_reg(reg_no)
        for other in students:
            if _clean_reg(other.get("reg_no")) != target_reg:
                continue
            alt_phone = _clean_phone(other.get("parent_phone")) or _clean_phone(other.get("parent_email"))
            if alt_phone:
                phone = alt_phone
                try:
                    db.update_student(email, reg_no, parent_phone=phone)
                except Exception:
                    pass
                break

    if not phone:
        flash(f"Parent phone number missing for {reg_no}.", "error")
        if is_ajax:
            return jsonify({"success": False, "error": f"Parent phone number missing for {normalized_reg}."}), 400
        return redirect(url_for("counselor_test_send_page", test_id=test_id))

    wa = get_whatsapp_link(phone, msg)
    db.log_message(email, normalized_reg, stu.get("student_name", ""), msg, "message", wa, test_id=test_id)

    if is_ajax:
        return jsonify({"success": True, "status": "generated", "wa_link": wa})

    # Open WhatsApp compose URL so counselor can send immediately.
    return redirect(wa)


@app.route("/api/reports/generate", methods=["POST"])
@login_required
def api_generate_reports():
    email = session["user_email"]
    test_id = request.form.get("test_id")
    reg_nos = request.form.getlist("reg_nos")
    fmt = request.form.get("format", "message")

    # Editable parsed fields
    edited_test_name = request.form.get("edited_test_name", "").strip()
    edited_semester = request.form.get("edited_semester", "").strip()
    edited_department = request.form.get("edited_department", "").strip()
    edited_batch = request.form.get("edited_batch", "").strip()
    custom_message_body = request.form.get("custom_message_body", "").strip()

    if not test_id or not reg_nos:
        flash("Select a test and at least one student.", "error")
        return redirect(url_for("counselor_page", tab="test-database", test_id=test_id or ""))

    test_id = int(test_id)
    user = db.get_user(email)
    test_meta = db.get_test_metadata(test_id)

    if _is_test_blocked(test_id):
        flash("This test is blocked by administration. Sending is disabled.", "warning")
        return redirect(url_for("counselor_page", tab="test-database", test_id=test_id))

    # Keep metadata editable before send
    if edited_test_name or edited_semester or edited_department or edited_batch:
        db.update_test_metadata_fields(
            test_id,
            test_name=edited_test_name or (test_meta or {}).get("test_name") or "",
            semester=edited_semester or (test_meta or {}).get("semester") or "",
            department=edited_department or (test_meta or {}).get("department") or "",
            batch_name=edited_batch or (test_meta or {}).get("batch_name") or "",
        )
        test_meta = db.get_test_metadata(test_id)

    students = db.get_students(email)
    lookup = {s["reg_no"]: s for s in students}

    from utils.whatsapp_helper import get_whatsapp_link
    from utils.template_engine import TemplateEngine

    already_sent = db.get_sent_reg_nos_for_test(email, test_id)
    reports = []
    for rn in reg_nos:
        if rn in already_sent:
            continue

        marks = db.get_student_marks_for_reg(test_id, rn)
        stu = lookup.get(rn, {})
        if not marks:
            continue

        marks_table = _build_parent_subjects_table(marks)
        test_name = (test_meta or {}).get("test_name") or "Unit Test"
        semester = (test_meta or {}).get("semester") or "-"
        department = (test_meta or {}).get("department") or stu.get("department", "-")
        batch_name = (test_meta or {}).get("batch_name") or "-"

        if custom_message_body:
            msg = TemplateEngine.fill_template(
                custom_message_body,
                app_name=APP_NAME,
                reg_no=rn,
                student_name=stu.get("student_name", rn),
                department=department,
                test_name=test_name,
                semester=semester,
                batch_name=batch_name,
                subjects_table=marks_table,
                counselor_name=user["name"],
            )
        else:
            msg = _build_parent_message(test_name, rn, stu.get('student_name', rn), marks)

        phone = stu.get("parent_phone", "")
        wa = get_whatsapp_link(phone, msg) if phone else ""
        db.log_message(email, rn, stu.get("student_name", ""), msg, fmt, wa, test_id=test_id)

        reports.append({
            "reg_no": rn,
            "name": stu.get("student_name", rn),
            "marks": marks,
            "message": msg,
            "whatsapp_link": wa,
            "phone": phone,
        })

    session["reports"] = reports
    session["report_test_id"] = test_id
    flash(f"Reports generated for {len(reports)} pending students.", "success")
    return redirect(url_for("counselor_page", tab="test-database", test_id=test_id))


@app.route("/api/reports/pdf/<reg_no>")
@login_required
def api_student_pdf(reg_no):
    email = session["user_email"]
    test_id = request.args.get("test_id")
    if not test_id:
        return "Missing test_id", 400

    marks = db.get_student_marks_for_reg(int(test_id), reg_no)
    if not marks:
        return "No marks found", 404

    stu = next((s for s in db.get_students(email) if s["reg_no"] == reg_no), None)
    user = db.get_user(email)
    meta = db.get_test_metadata(int(test_id))

    from utils.pdf_generator import generate_student_pdf
    pdf_bytes = generate_student_pdf(
        student_name=stu["student_name"] if stu else reg_no,
        reg_no=reg_no,
        department=(stu or {}).get("department", ""),
        subjects_marks=marks,
        counselor_name=user["name"],
        test_name=(meta or {}).get("test_name", "Unit Test"),
    )
    return send_file(
        io.BytesIO(pdf_bytes), mimetype="application/pdf",
        as_attachment=True, download_name=f"{reg_no}_report.pdf",
    )


# ---------- Format Settings -------------------------------------------------

@app.route("/api/settings/format", methods=["POST"])
@login_required
@admin_required
def api_update_format_settings():
    default = request.form.get("default_format", "message")
    allowed = request.form.getlist("allowed_formats")
    bulk = request.form.get("bulk_format", "same_as_individual")
    db.update_format_settings(default, allowed, bulk, session["user_email"])
    flash("Format settings updated.", "success")
    return _redirect_admin_back("config")


# ---------- Static data assets ----------------------------------------------

@app.route("/data/<path:filename>")
def serve_data(filename):
    return send_file(os.path.join(DATA_DIR, filename))


# ---------------------------------------------------------------------------
# Run
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    import sys
    port = int(sys.argv[1]) if len(sys.argv) > 1 else 5000
    print(f"\n  [OK] RMKCET Parent Connect running at: http://localhost:{port}\n")
    app.run(debug=False, host="0.0.0.0", port=port, use_reloader=False)
